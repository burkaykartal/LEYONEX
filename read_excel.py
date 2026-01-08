import openpyxl
import json
from datetime import datetime

# Excel dosyasını oku
wb = openpyxl.load_workbook('2026 ve sonrası fuar listesi.xlsx')
ws = wb.active

# Başlıkları al
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(str(cell.value).strip())

print('Başlıklar:', headers)
print('\n' + '='*80 + '\n')

# Verileri al
fairs = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[0]:  # İlk sütun boş değilse
        fair_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                if value is not None:
                    if isinstance(value, datetime):
                        fair_data[headers[i]] = value.strftime('%Y-%m-%d')
                    else:
                        fair_data[headers[i]] = str(value).strip()
                else:
                    fair_data[headers[i]] = ''
        fairs.append(fair_data)

# İlk 5 fuarı göster
print('İlk 5 fuar:')
for i, fair in enumerate(fairs[:5], 1):
    print(f'\nFuar {i}:')
    for key, value in fair.items():
        print(f'  {key}: {value}')
    print('-' * 80)

print(f'\nToplam fuar sayısı: {len(fairs)}')

# JSON dosyasına kaydet
with open('fairs_data.json', 'w', encoding='utf-8') as f:
    json.dump(fairs, f, ensure_ascii=False, indent=2)

print('\nVeriler fairs_data.json dosyasına kaydedildi.')
